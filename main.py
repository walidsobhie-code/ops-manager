"""
brain/analytics.py
──────────────────
Pure-Python analytics helpers for OPS NEXUS dashboard.
Core API has no external dependencies beyond the stdlib. 
Heavy components (Prophet, Scipy, openpyxl, WeasyPrint, OR-Tools) 
use dynamic imports to keep deployment fast and flexible.

Public API
----------
analyze_store_status(current_value, baseline)            -> str  "Green" | "Yellow" | "Red"
calculate_7day_baseline(store_reports)                   -> float
identify_red_zone_stores(today_reports, baselines)       -> list[dict]
generate_fleet_summary_prompt(recent_reports)            -> str
calculate_fleet_kpis(all_reports)                        -> dict
generate_store_forecast(store_reports, periods=7)        -> dict
calculate_store_benchmarks(all_reports)                  -> list[dict]
export_fleet_to_excel(df, output_path)                   -> str
export_fleet_to_pdf(df, output_path)                     -> str
optimize_staffing(store_id, req_hours, available_staff)  -> dict
"""

from __future__ import annotations

import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# 1. analyze_store_status
# ─────────────────────────────────────────────────────────────────────────────
def analyze_store_status(current_value: float, baseline: float) -> str:
    """
    Compare today's sales against the 7-day rolling baseline and return a
    traffic-light status string.

    Thresholds
    ----------
    Green  : current >= 90 % of baseline  (on target or above)
    Yellow : 70 % <= current < 90 %       (mild underperformance)
    Red    : current < 70 % of baseline   (significant drop)

    If baseline is 0 (new store / no history), returns "Green" to avoid
    false alarms.
    """
    if baseline <= 0:
        return "Green"

    ratio = current_value / baseline

    if ratio >= 0.90:
        return "Green"
    elif ratio >= 0.70:
        return "Yellow"
    else:
        return "Red"


# ─────────────────────────────────────────────────────────────────────────────
# 2. calculate_7day_baseline
# ─────────────────────────────────────────────────────────────────────────────
def calculate_7day_baseline(store_reports: List[Dict[str, Any]]) -> float:
    """
    Calculate a rolling 7-day average sales baseline for a single store.

    Parameters
    ----------
    store_reports : list of report dicts, each containing at minimum:
        - 'report_date'  : str (ISO "YYYY-MM-DD") or datetime / date object
        - 'sales'        : numeric or str-numeric

    Returns
    -------
    float  — average daily sales over the most recent 7 calendar days
             that have at least one report; 0.0 if no valid data.
    """
    if not store_reports:
        return 0.0

    # Normalise dates and sales
    parsed: List[tuple[datetime, float]] = []
    for r in store_reports:
        try:
            raw_date = r.get("report_date")
            if isinstance(raw_date, str):
                raw_date = datetime.fromisoformat(raw_date.split("T")[0])
            elif hasattr(raw_date, "date"):          # datetime / pandas Timestamp
                raw_date = raw_date                  # already datetime-like
            sales = float(r.get("sales") or 0)
            parsed.append((raw_date, sales))
        except (TypeError, ValueError):
            continue

    if not parsed:
        return 0.0

    # Find the most recent date in the dataset
    most_recent = max(d for d, _ in parsed)
    cutoff      = most_recent - timedelta(days=6)   # last 7 days inclusive

    recent_sales = [s for d, s in parsed if d >= cutoff and s > 0]

    if not recent_sales:
        # Fall back to all-time average if last 7 days are empty
        all_sales = [s for _, s in parsed if s > 0]
        return round(sum(all_sales) / len(all_sales), 2) if all_sales else 0.0

    return round(sum(recent_sales) / len(recent_sales), 2)


# ─────────────────────────────────────────────────────────────────────────────
# 3. identify_red_zone_stores
# ─────────────────────────────────────────────────────────────────────────────
def identify_red_zone_stores(
    today_reports: List[Dict[str, Any]],
    baselines: Dict[str, float],
    threshold_pct: float = 30.0,
) -> List[Dict[str, Any]]:
    """
    Identify stores whose current sales are below their baseline by at least
    `threshold_pct` percent.
    """
    red_zone: List[Dict[str, Any]] = []

    for report in today_reports:
        store_id = report.get("store_id")
        if not store_id:
            continue

        current = float(report.get("sales") or 0)
        baseline = baselines.get(store_id, 0.0)

        if baseline <= 0:
            continue

        drop_pct = (baseline - current) / baseline * 100

        if drop_pct >= threshold_pct:
            red_zone.append({
                "store_id":      store_id,
                "current_value": round(current, 2),
                "baseline":      round(baseline, 2),
                "drop_pct":      round(drop_pct, 1),
            })

    red_zone.sort(key=lambda x: x["drop_pct"], reverse=True)
    return red_zone


# ─────────────────────────────────────────────────────────────────────────────
# 4. generate_fleet_summary_prompt
# ─────────────────────────────────────────────────────────────────────────────
def generate_fleet_summary_prompt(recent_reports: List[Dict[str, Any]]) -> str:
    """
    Build a structured prompt for the Groq AI fleet intelligence banner.
    """
    if not recent_reports:
        return (
            'Return this exact JSON: {"fleet_health_score": 0, '
            '"strategic_recommendation": "No data available.", '
            '"critical_alerts": [], "top_performer": "", "at_risk_stores": []}'
        )

    store_lines: List[str] = []
    for r in recent_reports:
        store_id  = r.get("store_id", "Unknown")
        sales     = r.get("sales", 0)
        inventory = r.get("inventory_status", "unknown")
        staffing  = r.get("staffing", "unknown")
        analysis  = r.get("analysis", "")
        store_lines.append(
            f"  • {store_id}: sales=${sales}, inventory={inventory}, "
            f"staffing={staffing}, AI note={analysis}"
        )

    stores_block = "\n".join(store_lines)
    report_date  = recent_reports[0].get("report_date", "today")

    prompt = f"""You are the AI operations brain for a multi-store retail fleet.
Analyse the latest daily report data below and return a JSON object — nothing else.

REPORT DATE: {report_date}
STORES ({len(recent_reports)} total):
{stores_block}

Return ONLY valid JSON matching this exact schema (no markdown, no preamble):
{{
  "fleet_health_score": <integer 0-100, where 100 = all stores exceeding targets>,
  "strategic_recommendation": "<one clear, actionable sentence for the ops manager>",
  "critical_alerts": ["<alert 1>", "<alert 2>"],
  "top_performer": "<store_id of best performing store>",
  "at_risk_stores": ["<store_id>", ...]
}}

Scoring guide:
- 90-100 : All stores ≥ target, no inventory or staffing issues
- 70-89  : Minor shortfalls in 1-2 stores
- 50-69  : Multiple stores underperforming or staffing gaps
- 30-49  : Significant revenue decline across fleet
- 0-29   : Critical — multiple stores in crisis
"""
    return prompt


# ─────────────────────────────────────────────────────────────────────────────
# 5. calculate_fleet_kpis
# ─────────────────────────────────────────────────────────────────────────────
def calculate_fleet_kpis(all_reports: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Calculate top-level fleet KPIs from the full report dataset.
    """
    if not all_reports:
        return {
            "total_sales":  0.0,
            "avg_sales":    0.0,
            "store_count":  0,
            "report_count": 0,
            "top_store":    "—",
            "top_sales":    0.0,
        }

    store_totals: Dict[str, float] = {}
    for r in all_reports:
        sid   = r.get("store_id", "Unknown")
        sales = float(r.get("sales") or 0)
        store_totals[sid] = store_totals.get(sid, 0.0) + sales

    total_sales  = sum(store_totals.values())
    store_count  = len(store_totals)
    avg_sales    = round(total_sales / store_count, 2) if store_count else 0.0
    top_store    = max(store_totals, key=store_totals.get) if store_totals else "—"
    top_sales    = store_totals.get(top_store, 0.0)

    return {
        "total_sales":  round(total_sales, 2),
        "avg_sales":    avg_sales,
        "store_count":  store_count,
        "report_count": len(all_reports),
        "top_store":    top_store,
        "top_sales":    round(top_sales, 2),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 6. generate_store_forecast (Prophet)
# ─────────────────────────────────────────────────────────────────────────────
def generate_store_forecast(store_reports: List[Dict[str, Any]], periods: int = 7) -> Dict[str, Any]:
    """
    Predict future sales using Facebook Prophet. Safe dynamic import.
    """
    try:
        from prophet import Prophet
        import pandas as pd
    except ImportError:
        logger.warning("Prophet or Pandas not installed. Forecasting disabled.")
        return {"forecast": [], "trend": "library_missing"}
        
    if not store_reports:
        return {"forecast": [], "trend": "stable"}

    data = []
    for r in store_reports:
        date = r.get("report_date")
        if isinstance(date, str):
            date = date.split("T")[0]
        data.append({"ds": date, "y": float(r.get("sales") or 0)})
    
    df_prophet = pd.DataFrame(data)
    if len(df_prophet) < 2:
        return {"forecast": [], "trend": "insufficient_data"}

    model = Prophet(daily_seasonality=True, yearly_seasonality=False, weekly_seasonality=True)
    model.fit(df_prophet)
    
    future = model.make_future_dataframe(periods=periods)
    forecast = model.predict(future)
    
    result = forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(periods).to_dict('records')
    trend = "increasing" if result[-1]['yhat'] > result[0]['yhat'] else "decreasing"
    
    return {"forecast": result, "trend": trend}


# ─────────────────────────────────────────────────────────────────────────────
# 7. calculate_store_benchmarks (Z-Score)
# ─────────────────────────────────────────────────────────────────────────────
def calculate_store_benchmarks(all_reports: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Perform store-to-store benchmarking using Z-scores to find outliers. Safe dynamic import.
    """
    try:
        from scipy import stats
    except ImportError:
        logger.warning("Scipy not installed. Benchmarking disabled.")
        return []
        
    if not all_reports:
        return []

    latest_sales = {}
    for r in all_reports:
        sid = r.get("store_id")
        latest_sales[sid] = float(r.get("sales") or 0)
    
    sales_values = list(latest_sales.values())
    if len(sales_values) < 2:
        return []

    z_scores = stats.zscore(sales_values)
    
    benchmarks = []
    for i, (sid, val) in enumerate(latest_sales.items()):
        z = z_scores[i]
        significance = "Neutral"
        if z > 1.5: significance = "High Outlier (Positive)"
        elif z < -1.5: significance = "Critical Underperformer"
        
        benchmarks.append({
            "store_id": sid,
            "sales": val,
            "z_score": round(z, 2),
            "significance": significance
        })
    
    benchmarks.sort(key=lambda x: x['z_score'])
    return benchmarks


# ─────────────────────────────────────────────────────────────────────────────
# 8. export_to_excel
# ─────────────────────────────────────────────────────────────────────────────
def export_fleet_to_excel(df: Any, output_path: str = "fleet_report.xlsx") -> str:
    """
    Exports current fleet dataframe to a formatted Excel file. Safe dynamic import.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed. Excel export fallback to CSV.")
        csv_path = output_path.replace(".xlsx", ".csv")
        df.to_csv(csv_path, index=False)
        return csv_path

    df.to_excel(output_path, index=False, sheet_name="Ops Report")
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# 9. export_to_pdf
# ─────────────────────────────────────────────────────────────────────────────
def export_fleet_to_pdf(df: Any, output_path: str = "fleet_report.pdf") -> str:
    """
    Converts a simple HTML summary table to PDF. Safe dynamic import.
    """
    try:
        from weasyprint import HTML
    except ImportError:
        logger.warning("weasyprint not installed. PDF export skipped.")
        return ""

    html_content = f"""
    <html>
    <style>
        body {{ font-family: Arial, sans-serif; color: #333; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; }}
        h1 {{ color: #1f4e78; }}
    </style>
    <body>
        <h1>Sovereign Ops Fleet Report</h1>
        <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
        {df.to_html(index=False)}
    </body>
    </html>
    """
    HTML(string=html_content).write_pdf(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# 10. optimize_staffing (OR-Tools)
# ─────────────────────────────────────────────────────────────────────────────
def optimize_staffing(store_id: str, required_hours: int, available_staff: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Simple constraint solver to allocate shifts given available staff and total hour requirements.
    """
    try:
        from ortools.sat.python import cp_model
    except ImportError:
        logger.warning("ortools not installed. Staff optimization skipped.")
        return {"status": "library_missing", "schedule": []}

    model = cp_model.CpModel()
    
    staff_vars = {}
    for i, s in enumerate(available_staff):
        staff_vars[i] = model.NewIntVar(0, s.get("max_hours", 8), f"staff_{i}")
    
    model.Add(sum(staff_vars.values()) == required_hours)
    
    solver = cp_model.CpSolver()
    status = solver.Solve()
    
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        schedule = []
        for i, s in enumerate(available_staff):
            schedule.append({
                "name": s.get("name", f"Staff {i}"),
                "assigned_hours": solver.Value(staff_vars[i])
            })
        return {"status": "success", "schedule": schedule}
    
    return {"status": "infeasible", "schedule": []}
