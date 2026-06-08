from __future__ import annotations
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional
import pandas as pd
import numpy as np
from scipy import stats
from prophet import Prophet

logger = logging.getLogger(__name__)

def analyze_store_status(current_value: float, baseline: float) -> str:
    if baseline <= 0: return "Green"
    ratio = current_value / baseline
    if ratio >= 0.90: return "Green"
    elif ratio >= 0.70: return "Yellow"
    else: return "Red"

def calculate_7day_baseline(store_reports: List[Dict[str, Any]]) -> float:
    if not store_reports: return 0.0
    parsed = []
    for r in store_reports:
        try:
            raw_date = r.get("report_date")
            if isinstance(raw_date, str): raw_date = datetime.fromisoformat(raw_date.split("T")[0])
            elif hasattr(raw_date, "date"): raw_date = raw_date
            sales = float(r.get("sales") or 0)
            parsed.append((raw_date, sales))
        except: continue
    if not parsed: return 0.0
    most_recent = max(d for d, _ in parsed)
    cutoff = most_recent - timedelta(days=6)
    recent_sales = [s for d, s in parsed if d >= cutoff and s > 0]
    if not recent_sales:
        all_sales = [s for _, s in parsed if s > 0]
        return round(sum(all_sales) / len(all_sales), 2) if all_sales else 0.0
    return round(sum(recent_sales) / len(recent_sales), 2)

def identify_red_zone_stores(today_reports, baselines, threshold_pct=30.0):
    red_zone = []
    for report in today_reports:
        store_id = report.get("store_id")
        if not store_id: continue
        current = float(report.get("sales") or 0)
        baseline = baselines.get(store_id, 0.0)
        if baseline <= 0: continue
        drop_pct = (baseline - current) / baseline * 100
        if drop_pct >= threshold_pct:
            red_zone.append({"store_id": store_id, "current_value": round(current, 2), "baseline": round(baseline, 2), "drop_pct": round(drop_pct, 1)})
    red_zone.sort(key=lambda x: x["drop_pct"], reverse=True)
    return red_zone

def generate_fleet_summary_prompt(recent_reports):
    if not recent_reports: return 'Return this exact JSON: {"fleet_health_score": 0, "strategic_recommendation": "No data available.", "critical_alerts": [], "top_performer": "", "at_risk_stores": []}'
    store_lines = [f"  • {r.get('store_id', 'Unknown')}: sales=${r.get('sales', 0)}, inventory={r.get('inventory_status', 'unknown')}, staffing={r.get('staffing', 'unknown')}" for r in recent_reports]
    stores_block = "\n".join(store_lines)
    report_date = recent_reports[0].get("report_date", "today")
    return f"You are the AI operations brain. Analyse reports for {report_date}:\n{stores_block}\nReturn ONLY JSON: {{\"fleet_health_score\": int, \"strategic_recommendation\": str, \"critical_alerts\": [], \"top_performer\": str, \"at_risk_stores\": []}}"

def calculate_fleet_kpis(all_reports):
    if not all_reports: return {"total_sales": 0.0, "avg_sales": 0.0, "store_count": 0, "report_count": 0, "top_store": "—", "top_sales": 0.0}
    store_totals = {}
    for r in all_reports:
        sid = r.get("store_id", "Unknown")
        store_totals[sid] = store_totals.get(sid, 0.0) + float(r.get("sales") or 0)
    total_sales = sum(store_totals.values())
    store_count = len(store_totals)
    top_store = max(store_totals, key=store_totals.get) if store_totals else "—"
    return {"total_sales": round(total_sales, 2), "avg_sales": round(total_sales/store_count, 2) if store_count else 0.0, "store_count": store_count, "report_count": len(all_reports), "top_store": top_store, "top_sales": round(store_totals.get(top_store, 0), 2)}

def generate_store_forecast(store_reports, periods=7):
    if not store_reports: return {"forecast": [], "trend": "stable"}
    data = [{"ds": (r.get("report_date").split("T")[0] if isinstance(r.get("report_date"), str) else r.get("report_date")), "y": float(r.get("sales") or 0)} for r in store_reports]
    df_p = pd.DataFrame(data)
    if len(df_p) < 2: return {"forecast": [], "trend": "insufficient_data"}
    m = Prophet(daily_seasonality=True, yearly_seasonality=False, weekly_seasonality=True)
    m.fit(df_p)
    future = m.make_future_dataframe(periods=periods)
    forecast = m.predict(future)
    res = forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(periods).to_dict('records')
    trend = "increasing" if res[-1]['yhat'] > res[0]['yhat'] else "decreasing"
    return {"forecast": res, "trend": trend}

def calculate_store_benchmarks(all_reports):
    if not all_reports: return []
    latest_sales = {r.get("store_id"): float(r.get("sales") or 0) for r in all_reports if r.get("store_id")}
    vals = list(latest_sales.values())
    if len(vals) < 2: return []
    z_scores = stats.zscore(vals)
    bench = []
    for i, (sid, val) in enumerate(latest_sales.items()):
        z = z_scores[i]
        sign = "Neutral"
        if z > 1.5: sign = "High Outlier (Positive)"
        elif z < -1.5: sign = "Critical Underperformer"
        bench.append({"store_id": sid, "sales": val, "z_score": round(z, 2), "significance": sign})
    bench.sort(key=lambda x: x['z_score'])
    return bench

def export_fleet_to_excel(df, output_path="fleet_report.xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    df.to_excel(output_path, index=False)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    wb.save(output_path)
    return output_path

def export_fleet_to_pdf(df, output_path="fleet_report.pdf"):
    from weasyprint import HTML
    html = f"<html><body style='font-family:sans-serif;'><h1 style='color:#1f4e78;'>Sovereign Ops Fleet Report</h1><p>Generated: {datetime.now()}</p>{df.to_html(index=False)}</body></html>"
    HTML(string=html).write_pdf(output_path)
    return output_path

def optimize_staffing(store_id, required_hours, available_staff):
    from ortools.sat.python import cp_model
    model = cp_model.CpModel()
    staff_vars = {i: model.NewIntVar(0, s.get("max_hours", 8), f"s{i}") for i, s in enumerate(available_staff)}
    model.Add(sum(staff_vars.values()) == required_hours)
    solver = cp_model.CpSolver()
    if solver.Solve() == cp_model.OPTIMAL or solver.Solve() == cp_model.FEASIBLE:
        return {"status": "success", "schedule": [{"name": s.get("name"), "assigned_hours": solver.Value(staff_vars[i])} for i, s in enumerate(available_staff)]}
    return {"status": "infeasible", "schedule": []}
